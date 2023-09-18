import openpyxl


def ReadData(file, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_num, column=col_num).value


def WriteData(file, sheet_name, row_num, col_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=col_num).value = data
    workbook.save(file)


def RowCount(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row
